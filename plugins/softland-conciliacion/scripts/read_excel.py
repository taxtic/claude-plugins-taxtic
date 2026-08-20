"""Lectura estructural del Excel de conciliacion bancaria BCI, por PERFIL
de entrada (Fase 9).

No interpreta significado de negocio: solo extrae celdas, agrupa los bloques
repetidos de factura/monto/centro de costo, y separa filas fuera de alcance
(CARGO, fila Total, filas de control/saldo, filas vacias) de los movimientos
candidatos que pasaran a normalize.py.

Perfiles (rules/excel-layouts.json), auditados Fase 9 contra archivos reales:
- LEGACY_CON_ENCABEZADO: comportamiento historico (fila de encabezado + 11
  columnas fijas + bloques detectados por texto). NO representa ningun
  archivo real confirmado -- se mantiene por compatibilidad/regresion.
- SIN_ENCABEZADO_ORIGINAL: estructura real de PRUEBA CLAUDIO.xlsx (13
  columnas, sin encabezado, un bloque factura+monto).
- SIN_ENCABEZADO_COMPACTO: estructura real de CLAUDIO.xlsx (10 columnas,
  sin encabezado, sin saldo_contable ni cliente independiente).

Los 3 perfiles producen exactamente el mismo contrato canonico de campos
(ver 'campos_canonicos' en rules/excel-layouts.json) que normalize.py ya
consumia antes de esta fase -- ningun cambio aqui afecta a normalize.py,
validate.py, approval.py, transform.py, export_softland.py ni a
rules/softland-*.json.

Deteccion de perfil: nunca por nombre de archivo, nunca solo por cantidad
de columnas. Combina ausencia/presencia de texto de encabezado + forma de
fecha + forma de RUT en la posicion esperada de cada perfil sin encabezado.
Si la evidencia es insuficiente o ambigua: ESTRUCTURA_NO_RECONOCIDA
(LecturaError), nunca se elige un perfil por aproximacion ni se devuelve
silenciosamente 0 candidatos/0 omitidos para un archivo con datos.

El archivo de entrada se abre siempre en modo lectura y nunca se modifica.
"""
import argparse
import json
import re
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

RULES_PATH_DEFAULT = Path(__file__).resolve().parent.parent / "rules" / "excel-layouts.json"
FILA_ENCABEZADO_DEFAULT = 2
HOJA_DEFAULT = "Hoja1"


class LecturaError(ValueError):
    """Error explicito de lectura, con codigo estable para tests."""
    def __init__(self, codigo, mensaje):
        self.codigo = codigo
        super().__init__(f"{codigo}: {mensaje}")


def _cargar_reglas(path=None):
    p = Path(path) if path else RULES_PATH_DEFAULT
    with open(p, encoding="utf-8") as f:
        return json.load(f)


def _serializar(valor):
    if isinstance(valor, (datetime, date)):
        return valor.isoformat()
    return valor


def _convertir_fecha_serial(valor, epoch):
    """Si 'valor' llega como serial Excel crudo (int/float) -- caso real de
    una celda de fecha con number_format 'General' en vez de un formato de
    fecha, donde openpyxl no la reconoce automaticamente como datetime --
    lo convierte usando el epoch real del workbook. Aplica exclusivamente al
    campo 'fecha'. Nunca inventa una fecha: si la conversion falla, devuelve
    el valor original sin cambios para que el flujo existente (normalize.py/
    validate.py, sin modificar) termine en FECHA_NO_PARSEABLE/revision."""
    if isinstance(valor, bool):
        return valor
    if isinstance(valor, (int, float)):
        try:
            return from_excel(valor, epoch)
        except (ValueError, OverflowError, TypeError):
            return valor
    return valor


def _es_fecha_plausible(valor, rango_serial):
    if isinstance(valor, (datetime, date)):
        return True
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        return rango_serial[0] <= valor <= rango_serial[1]
    return False


def _es_rut_plausible(valor, patron_rut):
    if not isinstance(valor, str):
        return False
    texto = valor.strip().upper().replace(".", "")
    return bool(re.match(patron_rut, texto))


def _texto_celda_encabezado(ws, fila, columna):
    valor = ws.cell(row=fila, column=columna).value
    return valor.strip().upper() if isinstance(valor, str) else ""


def _fila_parece_encabezado(ws, perfil_legacy, fila_candidata):
    """Busca los tokens 'FECHA' y 'ABONO' en las columnas fijas de fecha y
    abonos del perfil LEGACY, en la fila candidata. No decide nada mas: si
    esto no coincide, se asume que el archivo no tiene fila de encabezado
    y se evalua contra los perfiles sin encabezado."""
    col_fecha = perfil_legacy["campos_fijos"]["fecha"]
    col_abonos = perfil_legacy["campos_fijos"]["abonos"]
    return (
        "FECHA" in _texto_celda_encabezado(ws, fila_candidata, col_fecha)
        and "ABONO" in _texto_celda_encabezado(ws, fila_candidata, col_abonos)
    )


def _detectar_perfil(ws, reglas):
    """Devuelve el nombre de perfil detectado, o lanza LecturaError con
    codigo ESTRUCTURA_NO_RECONOCIDA si la evidencia es insuficiente o
    ambigua. Nunca elige un perfil por aproximacion."""
    perfiles = reglas["perfiles"]
    legacy = perfiles["LEGACY_CON_ENCABEZADO"]
    fila_hdr = legacy.get("fila_encabezado_default", FILA_ENCABEZADO_DEFAULT)

    if _fila_parece_encabezado(ws, legacy, fila_hdr):
        return "LEGACY_CON_ENCABEZADO"

    # Sin fila de encabezado detectada: evaluar perfiles sin encabezado
    # contra la primera fila (asumida como primera fila de datos reales).
    fila_datos = 1
    deteccion = reglas["deteccion"]
    patron_rut = deteccion["patron_rut"]
    rango_serial = deteccion["rango_serial_fecha_excel"]

    perfiles_sin_encabezado = {
        nombre: cfg for nombre, cfg in perfiles.items()
        if not cfg.get("tiene_encabezado", False)
    }

    v_fecha = ws.cell(row=fila_datos, column=1).value
    if not _es_fecha_plausible(v_fecha, rango_serial):
        raise LecturaError(
            "ESTRUCTURA_NO_RECONOCIDA",
            f"La fila {fila_datos} no tiene fila de encabezado reconocible ni un valor de "
            f"fecha plausible en la columna 1 ({v_fecha!r}). No se puede determinar el perfil "
            f"de entrada sin adivinar -- revise el archivo manualmente.",
        )

    v_cargo = ws.cell(row=fila_datos, column=4).value
    if v_cargo is not None and not isinstance(v_cargo, (int, float)):
        raise LecturaError(
            "ESTRUCTURA_NO_RECONOCIDA",
            f"La columna 4 (posicion fija de 'cargo' en todos los perfiles) de la fila "
            f"{fila_datos} no es numerica ni vacia ({v_cargo!r}) -- no coincide con la "
            f"estructura de ningun perfil conocido.",
        )

    candidatos = []
    for nombre, cfg in perfiles_sin_encabezado.items():
        col_rut = cfg["campos_fijos"].get("rut_banco_crudo")
        if col_rut is None:
            continue
        v_rut = ws.cell(row=fila_datos, column=col_rut).value
        if _es_rut_plausible(v_rut, patron_rut):
            candidatos.append(nombre)

    if len(candidatos) == 1:
        return candidatos[0]
    if len(candidatos) > 1:
        raise LecturaError(
            "ESTRUCTURA_NO_RECONOCIDA",
            f"La fila {fila_datos} es compatible con mas de un perfil sin encabezado "
            f"({candidatos}) -- ambiguedad estructural, no se elige por aproximacion.",
        )
    raise LecturaError(
        "ESTRUCTURA_NO_RECONOCIDA",
        f"No se encontro un valor con forma de RUT valido en ninguna posicion esperada "
        f"de los perfiles sin encabezado, en la fila {fila_datos}. No se puede determinar "
        f"el perfil de entrada sin adivinar -- revise el archivo manualmente.",
    )


def _bloques_definicion_por_encabezado(headers, primera_columna_bloques):
    """Detecta bloques (factura, monto, centro) a partir del texto de encabezado,
    no de la letra de columna: el orden FACTURAS->MONTOS->CENTRO se repite pero
    el sufijo numerico de cada nombre varia (duplicados de Excel)."""
    bloques = []
    max_col = max(headers) if headers else 0
    c = primera_columna_bloques
    while c <= max_col:
        h = (headers.get(c) or "").strip().upper()
        if h.startswith("FACTURAS"):
            factura_col = c
            monto_col = None
            centro_col = None
            h_next = (headers.get(c + 1) or "").strip().upper()
            if h_next.startswith("MONTOS"):
                monto_col = c + 1
                h_next2 = (headers.get(c + 2) or "").strip().upper()
                if h_next2.startswith("CENTRO"):
                    centro_col = c + 2
            bloques.append({"factura_col": factura_col, "monto_col": monto_col, "centro_col": centro_col})
            c += 3 if centro_col else (2 if monto_col else 1)
        else:
            c += 1
    return bloques


def _columnas_reconocidas(campos_fijos, bloques_def):
    reconocidas = set(campos_fijos.values())
    for b in bloques_def:
        for k in ("factura_col", "monto_col", "centro_col"):
            if b[k]:
                reconocidas.add(b[k])
    return reconocidas


def _clasificar_fila(campos_fijos, tiene_bloques_con_datos):
    """Devuelve el motivo de exclusion (str) si la fila esta fuera de alcance
    para convertirse en Movimiento, o None si es candidata. Opera siempre
    sobre el contrato canonico de campos, independiente del perfil de
    entrada que los produjo."""
    fecha = campos_fijos.get("fecha")
    cargo = campos_fijos.get("cargo")
    abonos = campos_fijos.get("abonos")
    rut = campos_fijos.get("rut_banco_crudo")
    proveedor = campos_fijos.get("proveedor_cliente_banco_crudo")
    col_a_texto = str(fecha).strip().upper() if isinstance(fecha, str) else None

    if col_a_texto == "TOTAL":
        return "FILA_TOTAL"

    todo_vacio = all(
        campos_fijos.get(nombre) in (None, "")
        for nombre in CAMPOS_CANONICOS
    ) and not tiene_bloques_con_datos
    if todo_vacio:
        return "FILA_VACIA"

    if fecha is None and rut in (None, "") and proveedor in (None, ""):
        # sin fecha ni identificacion de cliente, pero con algun valor numerico
        # remanente (saldo/checksum) -> fila de control, no un movimiento real
        if any(campos_fijos.get(nombre) not in (None, "") for nombre in ("cargo", "abonos", "saldo_contable")):
            return "FILA_CONTROL_SALDO"
        if not tiene_bloques_con_datos:
            return "FILA_VACIA"

    if cargo not in (None, 0):
        return "CARGO_FUERA_DE_ALCANCE"

    return None


CAMPOS_CANONICOS = [
    "fecha", "n_cheque_transferencia", "n_correlativo", "cargo", "abonos",
    "saldo_contable", "rut_banco_crudo", "proveedor_cliente_banco_crudo",
    "cc", "detalle_transaccion", "cuenta_categoria_ingreso",
]


def leer_conciliacion(path, hoja=None, fila_encabezado=None, perfil=None, reglas=None):
    reglas = reglas if reglas is not None else _cargar_reglas()
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        return _leer_conciliacion_wb(wb, path, hoja, fila_encabezado, perfil, reglas)
    finally:
        # read_only=True mantiene un handle de archivo abierto hasta cerrar
        # explicitamente -- sin esto, en Windows el archivo queda bloqueado
        # (ej. tempfile.TemporaryDirectory no puede borrarlo despues).
        wb.close()


def _leer_conciliacion_wb(wb, path, hoja, fila_encabezado, perfil, reglas):
    nombre_hoja = hoja if hoja and hoja in wb.sheetnames else (
        HOJA_DEFAULT if HOJA_DEFAULT in wb.sheetnames else wb.sheetnames[0]
    )
    ws = wb[nombre_hoja]

    nombre_perfil = perfil or _detectar_perfil(ws, reglas)
    cfg_perfil = reglas["perfiles"].get(nombre_perfil)
    if cfg_perfil is None:
        raise LecturaError(
            "ESTRUCTURA_NO_RECONOCIDA",
            f"El perfil {nombre_perfil!r} no existe en rules/excel-layouts.json "
            f"(perfiles disponibles: {sorted(reglas['perfiles'].keys())}).",
        )

    campos_fijos_col = cfg_perfil["campos_fijos"]
    tiene_encabezado = cfg_perfil.get("tiene_encabezado", False)

    if tiene_encabezado:
        fila_hdr = fila_encabezado if fila_encabezado is not None else cfg_perfil.get(
            "fila_encabezado_default", FILA_ENCABEZADO_DEFAULT
        )
        fila_datos_inicio = fila_hdr + 1
        headers = {}
        for row in ws.iter_rows(min_row=fila_hdr, max_row=fila_hdr):
            for idx, cell in enumerate(row, start=1):
                headers[idx] = getattr(cell, "value", None)
        primera_columna_bloques = cfg_perfil.get("primera_columna_bloques", max(campos_fijos_col.values()) + 1)
        bloques_def = _bloques_definicion_por_encabezado(headers, primera_columna_bloques)
    else:
        fila_hdr = None
        fila_datos_inicio = 1
        bloques_def = [dict(b) for b in cfg_perfil.get("bloques_fijos", [])]

    reconocidas = _columnas_reconocidas(campos_fijos_col, bloques_def)

    primera_columna_extension = cfg_perfil.get(
        "primera_columna_bloques", (max(campos_fijos_col.values()) + 1) if campos_fijos_col else 1
    )
    columnas_no_reconocidas_headers = {}
    if tiene_encabezado:
        max_header_col = max(headers) if headers else 0
        columnas_no_reconocidas_headers = {
            get_column_letter(c): headers.get(c)
            for c in range(primera_columna_extension, max_header_col + 1)
            if c not in reconocidas and headers.get(c) not in (None, "")
        }

    movimientos_candidatos = []
    omitidos = []
    columnas_no_reconocidas_con_datos = set()

    fila_actual = fila_datos_inicio - 1
    for row in ws.iter_rows(min_row=fila_datos_inicio):
        fila_actual += 1
        celdas = {idx: getattr(cell, "value", None) for idx, cell in enumerate(row, start=1)}

        campos_fijos = {}
        for nombre in CAMPOS_CANONICOS:
            col = campos_fijos_col.get(nombre)
            valor = celdas.get(col) if col is not None else None
            if nombre == "fecha":
                valor = _convertir_fecha_serial(valor, wb.epoch)
            campos_fijos[nombre] = _serializar(valor)

        bloques = []
        for idx, b in enumerate(bloques_def, start=1):
            factura = _serializar(celdas.get(b["factura_col"]))
            monto = _serializar(celdas.get(b["monto_col"])) if b["monto_col"] else None
            centro = _serializar(celdas.get(b["centro_col"])) if b["centro_col"] else None
            if factura in (None, "") and monto in (None, "") and centro in (None, ""):
                continue
            bloques.append({
                "bloque_indice": idx,
                "factura": factura,
                "monto": monto,
                "centro_costo": centro,
            })

        columnas_no_reconocidas_fila = []
        for c, v in celdas.items():
            if c >= primera_columna_extension and c not in reconocidas and v not in (None, ""):
                letra = get_column_letter(c)
                columnas_no_reconocidas_con_datos.add(letra)
                columnas_no_reconocidas_fila.append({"columna": letra, "valor": _serializar(v)})

        motivo = _clasificar_fila(campos_fijos, tiene_bloques_con_datos=bool(bloques))
        registro = {
            "fila_origen": fila_actual,
            "hoja_origen": nombre_hoja,
            "campos_fijos": campos_fijos,
            "bloques": bloques,
            "columnas_no_reconocidas": columnas_no_reconocidas_fila,
        }
        if motivo:
            registro["motivo"] = motivo
            omitidos.append(registro)
        else:
            movimientos_candidatos.append(registro)

    return {
        "archivo_origen": str(path),
        "hoja_origen": nombre_hoja,
        "perfil_detectado": nombre_perfil,
        "fila_encabezado": fila_hdr,
        "movimientos_candidatos": movimientos_candidatos,
        "omitidos": omitidos,
        "columnas_no_reconocidas_headers": columnas_no_reconocidas_headers,
        "columnas_no_reconocidas_con_datos": sorted(columnas_no_reconocidas_con_datos),
    }


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("excel", help="Ruta al Excel de conciliacion (.xlsx)")
    parser.add_argument("--hoja", default=None, help="Nombre de hoja (default: Hoja1 o la primera disponible)")
    parser.add_argument("--fila-encabezado", type=int, default=None)
    parser.add_argument("--perfil", default=None, help="Fuerza un perfil (default: deteccion automatica)")
    parser.add_argument("--reglas", default=None, help="Ruta a excel-layouts.json (default: rules/excel-layouts.json del plugin)")
    parser.add_argument("--out", default=None, help="Ruta de salida JSON (default: stdout)")
    args = parser.parse_args(argv)

    reglas = _cargar_reglas(args.reglas)

    try:
        resultado = leer_conciliacion(
            Path(args.excel), hoja=args.hoja, fila_encabezado=args.fila_encabezado,
            perfil=args.perfil, reglas=reglas,
        )
    except LecturaError as e:
        print(f"Lectura abortada: {e}", file=sys.stderr)
        return 1

    salida = json.dumps(resultado, ensure_ascii=False, indent=2)

    if args.out:
        Path(args.out).write_text(salida, encoding="utf-8")
        print(f"Perfil detectado: {resultado['perfil_detectado']}")
        print(f"Movimientos candidatos: {len(resultado['movimientos_candidatos'])}")
        print(f"Omitidos: {len(resultado['omitidos'])}")
        if resultado["columnas_no_reconocidas_con_datos"]:
            print(f"AVISO: columnas no reconocidas con datos: {resultado['columnas_no_reconocidas_con_datos']}")
    else:
        print(salida)
    return 0


if __name__ == "__main__":
    sys.exit(main())
