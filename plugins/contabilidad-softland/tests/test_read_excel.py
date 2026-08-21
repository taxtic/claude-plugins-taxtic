import hashlib
import importlib.util
import os


def _load(name):
    spec = importlib.util.spec_from_file_location(
        name, os.path.join(os.path.dirname(__file__), "..", "scripts", name + ".py")
    )
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)
    return m


re = _load("read_excel")

FIXTURE = os.path.join(os.path.dirname(__file__), "fixtures", "conciliacion_fixture.xlsx")


def _leer():
    return re.leer_conciliacion(FIXTURE)


def _candidato_fila(resultado, fila):
    return next(c for c in resultado["movimientos_candidatos"] if c["fila_origen"] == fila)


def _omitido_fila(resultado, fila):
    return next(o for o in resultado["omitidos"] if o["fila_origen"] == fila)


def test_fila_normal_una_factura():
    resultado = _leer()
    c = _candidato_fila(resultado, 3)
    assert c["campos_fijos"]["abonos"] == 100000
    assert len(c["bloques"]) == 1
    assert c["bloques"][0]["factura"] == 1001
    assert c["bloques"][0]["monto"] == 100000


def test_fila_normal_multifactura():
    resultado = _leer()
    c = _candidato_fila(resultado, 4)
    assert len(c["bloques"]) == 2
    facturas = [b["factura"] for b in c["bloques"]]
    assert facturas == [2001, 2002]


def test_maximo_de_asignaciones_observado_en_fixture():
    resultado = _leer()
    max_bloques = max(len(c["bloques"]) for c in resultado["movimientos_candidatos"])
    assert max_bloques == 4
    c = _candidato_fila(resultado, 5)
    assert len(c["bloques"]) == 4


def test_fila_cargo_fuera_de_alcance():
    resultado = _leer()
    o = _omitido_fila(resultado, 12)
    assert o["motivo"] == "CARGO_FUERA_DE_ALCANCE"
    assert all(c["fila_origen"] != 12 for c in resultado["movimientos_candidatos"])


def test_fila_total_se_excluye():
    resultado = _leer()
    o = _omitido_fila(resultado, 13)
    assert o["motivo"] == "FILA_TOTAL"
    assert all(c["fila_origen"] != 13 for c in resultado["movimientos_candidatos"])


def test_fila_vacia_se_excluye():
    resultado = _leer()
    o = _omitido_fila(resultado, 14)
    assert o["motivo"] == "FILA_VACIA"
    assert all(c["fila_origen"] != 14 for c in resultado["movimientos_candidatos"])


def test_bloques_texto_no_factura_se_preservan_crudos():
    resultado = _leer()
    c = _candidato_fila(resultado, 7)
    assert c["bloques"][0]["factura"] == "DEVOLVER AL CLIENTE"
    c8 = _candidato_fila(resultado, 8)
    assert c8["bloques"][0]["factura"] == "AJUSTAR"
    c9 = _candidato_fila(resultado, 9)
    assert c9["bloques"][0]["factura"] == "BOLETA 81"


def test_na_en_proveedor_y_cuenta_no_excluye_la_fila():
    resultado = _leer()
    c = _candidato_fila(resultado, 10)
    assert c["campos_fijos"]["proveedor_cliente_banco_crudo"] == "#N/A"
    assert c["campos_fijos"]["cuenta_categoria_ingreso"] == "#N/A"
    assert c["campos_fijos"]["rut_banco_crudo"] == "83777777-7"


def test_multiples_rut_se_preservan_crudos_sin_dividir():
    resultado = _leer()
    c = _candidato_fila(resultado, 11)
    assert c["campos_fijos"]["rut_banco_crudo"] == "84888888-8 / 85999999-9"
    assert len(c["bloques"]) == 2


def test_no_modifica_el_excel_original():
    antes = hashlib.sha256(open(FIXTURE, "rb").read()).hexdigest()
    re.leer_conciliacion(FIXTURE)
    despues = hashlib.sha256(open(FIXTURE, "rb").read()).hexdigest()
    assert antes == despues


def test_deteccion_de_bloques_por_encabezado():
    resultado = _leer()
    # el fixture incluye una columna huerfana (X, "MONTOS999") equivalente a BD-BJ del Excel real
    assert resultado["columnas_no_reconocidas_headers"] == {"X": "MONTOS999"}
    assert resultado["columnas_no_reconocidas_con_datos"] == ["X"]


def test_columna_huerfana_no_se_convierte_en_bloque():
    """Equivalente estructural al hallazgo BD-BJ: una columna que empieza con
    'MONTOS' pero sin 'FACTURAS' precedente no debe agruparse en ningun bloque,
    y su dato no debe perderse silenciosamente."""
    resultado = _leer()
    c = _candidato_fila(resultado, 15)
    # el unico bloque reconocido en la fila es el de la factura 7001; la columna
    # huerfana X (valor 999) no genera un segundo bloque
    assert len(c["bloques"]) == 1
    assert c["bloques"][0]["factura"] == 7001
    # el valor de la columna huerfana se preserva con trazabilidad (columna + valor)
    assert c["columnas_no_reconocidas"] == [{"columna": "X", "valor": 999}]


# --- Fase 9: perfiles de entrada. Auditoria previa confirmo que ni
# PRUEBA CLAUDIO.xlsx ni CLAUDIO.xlsx (los unicos archivos reales
# disponibles) tienen fila de encabezado -- LEGACY_CON_ENCABEZADO se
# mantiene solo por compatibilidad con el fixture historico de arriba
# (nunca representa un archivo real confirmado). Todos los tests de esta
# seccion usan fixtures 100% ficticios. ---

FIXTURE_ORIGINAL = os.path.join(os.path.dirname(__file__), "fixtures", "conciliacion_fixture_regresion_sin_encabezado.xlsx")
FIXTURE_COMPACTA = os.path.join(os.path.dirname(__file__), "fixtures", "conciliacion_fixture_compacta.xlsx")


def test_legacy_sigue_siendo_el_perfil_detectado_para_el_fixture_historico():
    resultado = _leer()
    assert resultado["perfil_detectado"] == "LEGACY_CON_ENCABEZADO"
    assert resultado["fila_encabezado"] == 2


def test_sin_encabezado_original_se_detecta_automaticamente():
    resultado = re.leer_conciliacion(FIXTURE_ORIGINAL)
    assert resultado["perfil_detectado"] == "SIN_ENCABEZADO_ORIGINAL"
    assert resultado["fila_encabezado"] is None
    assert len(resultado["movimientos_candidatos"]) == 1


def test_sin_encabezado_original_no_pierde_factura_ni_monto():
    resultado = re.leer_conciliacion(FIXTURE_ORIGINAL)
    c = resultado["movimientos_candidatos"][0]
    assert len(c["bloques"]) == 1
    assert c["bloques"][0]["factura"] == 90001
    assert c["bloques"][0]["monto"] == 88000
    assert c["columnas_no_reconocidas"] == []


def test_sin_encabezado_compacto_se_detecta_automaticamente():
    resultado = re.leer_conciliacion(FIXTURE_COMPACTA)
    assert resultado["perfil_detectado"] == "SIN_ENCABEZADO_COMPACTO"
    assert len(resultado["movimientos_candidatos"]) == 1


def test_sin_encabezado_compacto_lee_los_campos_disponibles():
    resultado = re.leer_conciliacion(FIXTURE_COMPACTA)
    c = resultado["movimientos_candidatos"][0]["campos_fijos"]
    assert c["fecha"] == "2026-10-03T00:00:00"
    assert c["n_cheque_transferencia"] == "Transferencia recibida de EMPRESA FICTICIA DOS SPA"
    assert c["n_correlativo"] == 777
    assert c["cargo"] == 0
    assert c["abonos"] == 45500
    assert c["rut_banco_crudo"] == "98765432-1"
    assert c["detalle_transaccion"] == "PRUEBA COMPACTA REGRESION"


def test_sin_encabezado_compacto_campos_no_disponibles_son_none():
    """Contrato explicito (Fase 9): el perfil compacto NUNCA inventa un
    valor para los campos que su Excel real no trae -- los entrega en
    None, nunca los omite ni los infiere desde otro campo."""
    resultado = re.leer_conciliacion(FIXTURE_COMPACTA)
    c = resultado["movimientos_candidatos"][0]["campos_fijos"]
    assert c["saldo_contable"] is None
    assert c["proveedor_cliente_banco_crudo"] is None
    assert c["cc"] is None
    assert c["cuenta_categoria_ingreso"] is None


def test_sin_encabezado_compacto_no_infiere_nombre_cliente_de_la_descripcion():
    """PENDIENTE FUNCIONAL (Fase 9): aunque n_cheque_transferencia contenga
    un nombre que parezca una empresa, jamas se usa para completar
    proveedor_cliente_banco_crudo. La auditoria previa demostro con el
    propio caso real ya validado en Softland que esos dos campos pueden
    referirse a entidades distintas (persona que transfiere vs cliente
    contable)."""
    resultado = re.leer_conciliacion(FIXTURE_COMPACTA)
    c = resultado["movimientos_candidatos"][0]["campos_fijos"]
    assert "EMPRESA FICTICIA DOS SPA" in c["n_cheque_transferencia"]
    assert c["proveedor_cliente_banco_crudo"] is None


def test_sin_encabezado_compacto_bloque_no_desaparece():
    resultado = re.leer_conciliacion(FIXTURE_COMPACTA)
    c = resultado["movimientos_candidatos"][0]
    assert len(c["bloques"]) == 1
    assert c["bloques"][0]["factura"] == 80002
    assert c["bloques"][0]["monto"] == 45500


def test_estructura_no_reconocida_sin_fecha_plausible_ni_encabezado(tmp_path):
    """Nunca debe devolver 0 candidatos/0 omitidos silenciosamente para un
    archivo que si tiene datos: debe fallar explicitamente."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["texto sin forma de fecha", "algo", 1, 0, 1000, "no es un rut", "detalle", None, 123, 1000])
    ruta = tmp_path / "archivo_no_reconocido.xlsx"
    wb.save(ruta)

    try:
        re.leer_conciliacion(str(ruta))
        assert False, "debia fallar con ESTRUCTURA_NO_RECONOCIDA"
    except re.LecturaError as e:
        assert e.codigo == "ESTRUCTURA_NO_RECONOCIDA"


def test_estructura_no_reconocida_nunca_es_cero_candidatos_cero_omitidos():
    """Regresion explicita del hallazgo de la auditoria previa: correr
    read_excel.py contra un archivo real sin adaptar producia 0/0 en
    silencio. Ahora debe lanzar LecturaError en vez de eso."""
    import openpyxl
    import tempfile
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([99999999, "sin fecha ni rut reconocible", 1, 0, 1000, "ABC", "x", None, 1, 1000])
    with tempfile.TemporaryDirectory() as d:
        ruta = os.path.join(d, "archivo.xlsx")
        wb.save(ruta)
        try:
            re.leer_conciliacion(ruta)
            assert False, "debia fallar explicitamente, no devolver 0/0"
        except re.LecturaError as e:
            assert e.codigo == "ESTRUCTURA_NO_RECONOCIDA"


def test_estructura_ambigua_entre_perfiles_sin_encabezado_falla(tmp_path):
    """Si una fila fuera compatible con la forma de RUT esperada en MAS DE
    UN perfil sin encabezado simultaneamente, no se elige por aproximacion:
    se declara ambiguedad explicita."""
    import openpyxl
    from datetime import datetime
    wb = openpyxl.Workbook()
    ws = wb.active
    # columna 6 (RUT del perfil compacto) Y columna 7 (RUT del perfil original)
    # ambas con forma de RUT valida -> ambiguo
    ws.append([
        datetime(2026, 1, 1), "desc", 1, 0, 1000,
        "11111111-1",  # columna 6: forma de RUT (coincide con SIN_ENCABEZADO_COMPACTO)
        "22222222-2",  # columna 7: forma de RUT (coincide con SIN_ENCABEZADO_ORIGINAL)
        "cliente", None, "detalle", "categoria", 1, 1000,
    ])
    ruta = tmp_path / "archivo_ambiguo.xlsx"
    wb.save(ruta)

    try:
        re.leer_conciliacion(str(ruta))
        assert False, "debia fallar con ESTRUCTURA_NO_RECONOCIDA por ambiguedad"
    except re.LecturaError as e:
        assert e.codigo == "ESTRUCTURA_NO_RECONOCIDA"


def test_perfil_se_puede_forzar_explicitamente():
    """Permite a un llamador (o a un test) omitir la deteccion automatica
    cuando ya se sabe cual perfil corresponde."""
    resultado = re.leer_conciliacion(FIXTURE_ORIGINAL, perfil="SIN_ENCABEZADO_ORIGINAL")
    assert resultado["perfil_detectado"] == "SIN_ENCABEZADO_ORIGINAL"


def test_perfil_inexistente_falla_explicito():
    try:
        re.leer_conciliacion(FIXTURE_ORIGINAL, perfil="INVENTADO_99")
        assert False, "debia fallar"
    except re.LecturaError as e:
        assert e.codigo == "ESTRUCTURA_NO_RECONOCIDA"


def test_no_modifica_los_excel_reales_ficticios():
    for ruta in (FIXTURE_ORIGINAL, FIXTURE_COMPACTA):
        antes = hashlib.sha256(open(ruta, "rb").read()).hexdigest()
        re.leer_conciliacion(ruta)
        despues = hashlib.sha256(open(ruta, "rb").read()).hexdigest()
        assert antes == despues


# --- fecha como serial Excel crudo (number_format "General") -- caso real
# encontrado en PRUEBA CLAUDIO.xlsx: openpyxl no reconoce automaticamente la
# celda de fecha como datetime cuando el formato de la celda es "General",
# por lo que campos_fijos["fecha"] llega como int/float crudo. La conversion
# se hace en read_excel.py (openpyxl.utils.datetime.from_excel + wb.epoch),
# exclusivamente para el campo 'fecha', para que normalize.py (sin cambios)
# reciba la misma representacion que ya sabe consumir.

FIXTURE_FECHA_SERIAL = os.path.join(os.path.dirname(__file__), "fixtures", "conciliacion_fixture_fecha_serial.xlsx")


def test_fecha_serial_excel_crudo_se_convierte_correctamente():
    resultado = re.leer_conciliacion(FIXTURE_FECHA_SERIAL)
    c = resultado["movimientos_candidatos"][0]
    assert c["campos_fijos"]["fecha"] == "2026-08-07T00:00:00"


def test_fecha_serial_no_afecta_otros_campos_numericos():
    """La conversion es exclusiva del campo 'fecha' -- otros campos numericos
    del mismo fixture (n_correlativo, cargo, abonos, saldo_contable, factura,
    monto) deben permanecer como numeros, sin convertirse en fecha."""
    resultado = re.leer_conciliacion(FIXTURE_FECHA_SERIAL)
    c = resultado["movimientos_candidatos"][0]
    assert c["campos_fijos"]["n_correlativo"] == 601
    assert c["campos_fijos"]["cargo"] == 0
    assert c["campos_fijos"]["abonos"] == 73000
    assert c["campos_fijos"]["saldo_contable"] == 850000
    assert c["bloques"][0]["factura"] == 70003
    assert c["bloques"][0]["monto"] == 73000


def test_no_modifica_el_fixture_de_fecha_serial():
    antes = hashlib.sha256(open(FIXTURE_FECHA_SERIAL, "rb").read()).hexdigest()
    re.leer_conciliacion(FIXTURE_FECHA_SERIAL)
    despues = hashlib.sha256(open(FIXTURE_FECHA_SERIAL, "rb").read()).hexdigest()
    assert antes == despues
